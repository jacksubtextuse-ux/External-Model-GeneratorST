from openpyxl import load_workbook
p = r"C:\Users\JackBranding\OneDrive - Subtext\Desktop\Codex\workflows\verve-proforma-cleaner\samples\Proforma_Limestone_20260310MK_lexington_temp.xlsm"
wb = load_workbook(p, keep_vba=True)
ws = wb["Executive Summary"]
ws["E6"] = "100 Main St, Lexington KY, KY 40502"
wb.save(p)
print(p)
