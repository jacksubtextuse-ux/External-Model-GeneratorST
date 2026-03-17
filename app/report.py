from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DELETION_TARGETS = [
    "Building Program",
    "Construction Pricing",
    "3-Yr Co-GP Waterfall",
    "Reassessed 1-Yr Waterfall",
    "Reassessed 3-Yr Waterfall",
    "Reassessed 4-Yr Waterfall",
    "Historic OpEx Comparison",
    "Change Log",
    "TermSheet",
    "TermSheet-Ignore",
    "Proforma Comparison",
]

SPOT_CELLS = [
    "Executive Summary!J15",
    "Executive Summary!K15",
    "Executive Summary!L15",
    "Cash Flow!Q46",
    "Assumptions!I43",
    "Assumptions!X22",
    "Assumptions!D21",
    "Development!H11",
    "Development!H78",
    "Sale Proceeds!B75",
    "Returns Exhibit!J3",
]


def _split_cell(ref: str) -> tuple[str, str]:
    sheet, cell = ref.split("!", 1)
    return sheet, cell


def _cell_value(wb, ref: str) -> Any:
    sheet, cell = _split_cell(ref)
    if sheet not in wb.sheetnames:
        return None
    return wb[sheet][cell].value


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _step40_proof(wb) -> dict[str, Any]:
    if "Assumptions" not in wb.sheetnames:
        return {"ok": False, "error": "Assumptions sheet missing"}
    ws = wb["Assumptions"]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "residential parking spaces" in v.lower():
                target = ws.cell(r, c + 1)
                tv = target.value
                return {
                    "ok": not _is_formula(tv),
                    "label_cell": ws.cell(r, c).coordinate,
                    "label_text": v,
                    "target_cell": target.coordinate,
                    "target_value": tv,
                    "target_is_formula": _is_formula(tv),
                }
    return {"ok": False, "error": "Residential Parking Spaces label not found"}


def build_side_by_side_report(input_file: Path, output_file: Path) -> dict[str, Any]:
    in_wb = load_workbook(input_file, data_only=False, keep_vba=True)
    out_wb = load_workbook(output_file, data_only=False, keep_vba=True)

    in_sheets = set(in_wb.sheetnames)
    out_sheets = set(out_wb.sheetnames)
    removed = sorted(in_sheets - out_sheets)
    added = sorted(out_sheets - in_sheets)

    spot = []
    for ref in SPOT_CELLS:
        before = _cell_value(in_wb, ref)
        after = _cell_value(out_wb, ref)
        spot.append(
            {
                "cell": ref,
                "before": before,
                "after": after,
                "before_is_formula": _is_formula(before),
                "after_is_formula": _is_formula(after),
            }
        )

    expected_removed_still_present = sorted([s for s in DELETION_TARGETS if s in out_sheets])

    # Lightweight summary counts for changed formulas/literals on common sheets.
    formula_changes = 0
    literal_changes = 0
    for s in sorted(in_sheets & out_sheets):
        iws = in_wb[s]
        ows = out_wb[s]
        max_row = max(iws.max_row, ows.max_row)
        max_col = max(iws.max_column, ows.max_column)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                iv = iws.cell(r, c).value
                ov = ows.cell(r, c).value
                if _is_formula(iv) or _is_formula(ov):
                    if iv != ov:
                        formula_changes += 1
                else:
                    if iv != ov:
                        literal_changes += 1

    return {
        "input_file": str(input_file),
        "output_file": str(output_file),
        "sheet_diff": {
            "removed": removed,
            "added": added,
            "expected_removed_still_present": expected_removed_still_present,
        },
        "changes": {
            "formula_changes": formula_changes,
            "literal_changes": literal_changes,
        },
        "spot_checks": spot,
        "step40_proof": _step40_proof(out_wb),
    }
