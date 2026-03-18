from __future__ import annotations

import datetime as dt
import re
import os
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.formula.translate import Translator

WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
NO_BORDER = Border(
    left=Side(style=None),
    right=Side(style=None),
    top=Side(style=None),
    bottom=Side(style=None),
    diagonal=Side(style=None),
    diagonalDown=False,
    diagonalUp=False,
    outline=False,
    vertical=Side(style=None),
    horizontal=Side(style=None),
)


@dataclass
class RunLog:
    messages: list[str] = field(default_factory=list)

    def add(self, message: str) -> None:
        self.messages.append(message)


class WorkflowError(RuntimeError):
    pass


class VerveWorkflowRunner:
    def __init__(self, input_file: Path, option: str = "base"):
        self.input_file = Path(input_file)
        self.wb = load_workbook(self.input_file, data_only=False, keep_vba=True)
        self.values_wb = load_workbook(self.input_file, data_only=True, keep_vba=True)
        self.log = RunLog()
        self.option = option

    def run(self, output_dir: Path | None = None) -> dict[str, Any]:
        if self.option != "base":
            raise WorkflowError("Selected option requires COM runner; openpyxl fallback supports only base")
        output_dir = Path(output_dir) if output_dir else self.input_file.parent
        output_dir.mkdir(parents=True, exist_ok=True)
        city_compact, city_spaced = self._parse_city_from_e6()
        project_type = self._project_type_prefix()
        baseline_q46 = self._value("Cash Flow", "Q46")

        self._step_2_blue_to_black("Executive Summary", {"0066FF", "0070C0"})
        self._step_3_set_project_name(city_spaced, project_type)
        self._step_4_hardcode_cells("Executive Summary", ["J15", "K15", "L15"])
        self._step_5_dscr_check()
        self._step_6_delete_zero_opex_rows()
        self._step_7_clear_n1_u12_preserve_n_border()
        self._step_8_conditional_row_collapse()

        self._step_9_group_hide_cash_flow_rows()
        self._step_10_conditional_delete_cashflow_row()

        self._step_11_blue_to_black_assumptions()
        self._step_12_remove_comments("Assumptions")
        self._step_13_group_hide_assumptions_dash_rows()
        self._step_14_hardcode_retail_tbd_sf()
        self._step_15_hardcode_assumptions_external_refs()

        self._step_12_remove_comments("Development")
        self._step_17_non_black_non_white_to_black_development()
        self._step_18_hardcode_development_external_refs()
        self._step_19_group_hide_development_unused_rows()
        self._step_19b_clear_development_comments_description_values()

        self._step_20_group_hide_sale_proceeds_commercial()
        self._clear_to_white("Sale Proceeds", "B75:J112")

        self._hardcode_refs_to_sheet("Building Program")
        self._safe_delete_sheet("Building Program")
        self._hardcode_refs_to_sheet("Construction Pricing")
        self._safe_delete_sheet("Construction Pricing")
        self._hardcode_refs_to_sheet("3-Yr Co-GP Waterfall")
        self._safe_delete_sheet("3-Yr Co-GP Waterfall")
        self._step_28_clear_returns_exhibit()

        for s in [
            "Reassessed 1-Yr Waterfall",
            "Reassessed 3-Yr Waterfall",
            "Reassessed 4-Yr Waterfall",
            "Historic OpEx Comparison",
            "Change Log",
            "TermSheet",
            "TermSheet-Ignore",
            "Proforma Comparison",
        ]:
            self._safe_delete_sheet(s)

        self._step_36b_delete_user_requested_tabs()

        self._step_37_clear_development_range()
        self._step_38_clear_yellow_reference_cells()
        self._step_39_remove_non_approved_fill_colors_assumptions()
        self._step_40_hardcode_residential_parking_spaces()
        self._step_40b_hardcode_residential_parking_rent_stall()
        self._step_41_copy_assumptions_range()
        self._step_42_clear_assumptions_range()
        self._step_15b_clear_total_interim_income_adjacent()
        self._step_42b_remove_waterfall_comments_notes()

        self._assert_q46_consistency_best_effort(baseline_q46)
        today = dt.datetime.now().strftime("%Y%m%d")
        market_slug = self._market_slug(default=city_compact)
        out_name = f"{project_type}_{market_slug}_{today}.xlsm"
        out_path = output_dir / out_name
        self.wb.save(out_path)
        self.log.add(f"Step 1: file renamed on output to {out_name}")
        return {"output_file": str(out_path), "log": self.log.messages}

    def _sheet(self, name: str):
        for s in self.wb.worksheets:
            if s.title.lower() == name.lower():
                return s
        return None

    def _values_sheet(self, name: str):
        for s in self.values_wb.worksheets:
            if s.title.lower() == name.lower():
                return s
        return None

    def _value(self, sheet: str, cell: str) -> Any:
        ws = self._values_sheet(sheet)
        if not ws:
            return None
        return ws[cell].value

    def _assert_q46_consistency_best_effort(self, baseline_q46: Any) -> None:
        ws = self._sheet("Cash Flow")
        if ws is None:
            raise WorkflowError("Q46 integrity check failed: Cash Flow sheet missing")
        cell = ws["Q46"].value
        if isinstance(cell, str) and cell.startswith("="):
            # openpyxl cannot recalc formulas; keep as warning-only in fallback mode
            self.log.add("Q46 integrity check skipped in openpyxl mode (formula not recalculated)")
            return
        try:
            b = float(baseline_q46)
            f = float(cell)
        except Exception:
            raise WorkflowError(f"Q46 integrity check failed: non-numeric baseline/final values baseline={baseline_q46!r}, final={cell!r}")
        if abs(b - f) > 1e-8:
            raise WorkflowError(f"Q46 integrity check failed: baseline={b} final={f} (difference={f-b})")
        self.log.add(f"Q46 integrity check passed: baseline={b} final={f}")

    def _parse_city_from_e6(self) -> tuple[str, str]:
        v = self._value("Executive Summary", "E6")
        if not isinstance(v, str) or "," not in v:
            raise WorkflowError("Executive Summary!E6 missing parseable address")
        parts = [p.strip() for p in v.split(",")]
        if len(parts) < 2:
            raise WorkflowError("Executive Summary!E6 address parse failed")
        city_spaced = parts[1]
        city_compact = city_spaced.replace(" ", "")
        return city_compact, city_spaced


    def _tax_abatement_pref(self) -> str | None:
        raw = os.environ.get("VERVE_TAX_ABATEMENT", "").strip().lower()
        if raw in {"yes", "no"}:
            return raw
        return None

    def _project_type_prefix(self) -> str:
        raw = os.environ.get("VERVE_PROJECT_TYPE", "VERVE").strip().upper()
        if raw in {"VERVE", "EVER", "LOCAL"}:
            return raw
        return "VERVE"

    def _market_slug(self, default: str) -> str:
        raw = os.environ.get("VERVE_MARKET", "").strip().upper()
        if not raw:
            raw = default.strip().upper()
        raw = re.sub(r"[^A-Z0-9]+", "_", raw)
        raw = re.sub(r"_+", "_", raw).strip("_")
        return raw or "MARKET"

    def _protected_sheet_names_for_user_delete(self) -> set[str]:
        return {
            "executive summary",
            "development summary",
            "cash flow",
            "assumptions",
            "development",
            "sale proceeds",
            "returns exhibit",
            "1-yr waterfall",
            "3-yr waterfall",
            "4-yr waterfall",
        }

    def _additional_delete_tabs_from_env(self) -> list[str]:
        raw = os.environ.get("VERVE_ADDITIONAL_DELETE_TABS", "").strip()
        if not raw:
            return []

        names: list[str] = []
        seen: set[str] = set()

        parsed: Any = None
        try:
            import json

            parsed = json.loads(raw)
        except Exception:
            parsed = None

        if isinstance(parsed, list):
            items = parsed
        else:
            items = re.split(r"[\n,]+", raw)

        for item in items:
            name = str(item or "").strip()
            if not name:
                continue
            key = name.lower()
            if key in seen:
                continue
            names.append(name)
            seen.add(key)
        return names

    def _step_36b_delete_user_requested_tabs(self) -> None:
        tabs = self._additional_delete_tabs_from_env()
        if not tabs:
            self.log.add("Step 36b: no user-requested additional tabs")
            return

        protected = self._protected_sheet_names_for_user_delete()
        deleted: list[str] = []
        skipped_protected: list[str] = []
        for tab in tabs:
            if tab.lower() in protected:
                skipped_protected.append(tab)
                self.log.add(f"Step 36b: skipped protected sheet '{tab}'")
                continue
            self._hardcode_refs_to_sheet(tab)
            self._safe_delete_sheet(tab)
            deleted.append(tab)

        msg = f"Step 36b: processed {len(tabs)} user-requested tabs"
        if deleted:
            msg += f"; attempted delete/hardcode for: {', '.join(deleted)}"
        if skipped_protected:
            msg += f"; protected skips: {', '.join(skipped_protected)}"
        self.log.add(msg)
    def _normalize_rgb(self, rgb: Any) -> str | None:
        if not rgb:
            return None
        raw = str(rgb).upper()
        if len(raw) == 8:
            raw = raw[2:]
        if len(raw) == 6:
            return raw
        return None

    def _step_2_blue_to_black(self, sheet: str, targets: set[str]) -> None:
        ws = self._sheet(sheet)
        if not ws:
            self.log.add(f"Step 2 skipped: sheet '{sheet}' missing")
            return
        changed = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                color = self._normalize_rgb(getattr(c.font.color, "rgb", None) if c.font and c.font.color else None)
                if color in targets:
                    c.font = copy(c.font)
                    c.font = c.font.copy(color="000000")
                    changed += 1
        self.log.add(f"Step 2: updated {changed} cells on {sheet}")

    def _step_3_set_project_name(self, city_spaced: str, project_type: str) -> None:
        ws = self._sheet("Executive Summary")
        if ws:
            ws["E5"].value = f"{project_type} {city_spaced}"
            self.log.add("Step 3: set Executive Summary!E5")

    def _step_4_hardcode_cells(self, sheet: str, cells: list[str]) -> None:
        ws = self._sheet(sheet)
        if not ws:
            return
        for addr in cells:
            ws[addr].value = self._value(sheet, addr)
        self.log.add(f"Step 4: hardcoded {sheet} {', '.join(cells)}")

    def _step_5_dscr_check(self) -> None:
        ws = self._sheet("Executive Summary")
        v = self._value("Executive Summary", "E60")
        if ws and isinstance(v, (int, float)) and v < 1.25:
            ws["E60"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
            self.log.add(f"Step 5 warning: DSCR {v} below 1.25")
        else:
            self.log.add("Step 5: DSCR pass or unavailable")

    def _find_label_row(
        self,
        ws,
        column: str,
        label: str,
        min_row: int = 1,
        max_row: int | None = None,
        exact: bool = False,
    ) -> int | None:
        col_idx = column_index_from_string(column)
        max_row = max_row or ws.max_row
        want = label.strip().lower()
        for r in range(min_row, max_row + 1):
            val = ws.cell(r, col_idx).value
            if not isinstance(val, str):
                continue
            have = val.strip().lower()
            if (exact and have == want) or ((not exact) and want in have):
                return r
        return None
    def _step_6_delete_zero_opex_rows(self) -> None:
        ws = self._sheet("Executive Summary")
        vws = self._values_sheet("Executive Summary")
        if not ws or not vws:
            return

        labels = ["Ground Lease", "Tax Abatement"]
        tax_pref = self._tax_abatement_pref()
        deleted_above = 0
        for label in labels:
            val_row = self._find_label_row(vws, "G", label, min_row=50, max_row=80, exact=True)
            if not val_row:
                self.log.add(f"Step 6: label '{label}' not found in OpEx section")
                continue

            val = vws[f"K{val_row}"].value
            target_row = val_row - deleted_above

            force_delete = label == "Tax Abatement" and tax_pref == "no"
            force_keep = label == "Tax Abatement" and tax_pref == "yes"
            should_delete = force_delete or (not force_keep and isinstance(val, (int, float)) and val == 0)

            if should_delete:
                ws.move_range(f"G{target_row + 1}:L{ws.max_row}", rows=-1, cols=0, translate=True)
                for col in range(column_index_from_string("G"), column_index_from_string("L") + 1):
                    ws.cell(ws.max_row, col).value = None
                    ws.cell(ws.max_row, col).border = copy(NO_BORDER)
                    ws.cell(ws.max_row, col).fill = copy(WHITE_FILL)
                deleted_above += 1
                mode = "forced by tax_abatement=no" if force_delete else "value==0"
                self.log.add(f"Step 6: deleted {label} row {target_row} ({mode})")
            else:
                mode = "forced keep by tax_abatement=yes" if force_keep else f"value {val}"
                self.log.add(f"Step 6: kept {label} row {target_row} ({mode})")
    def _clear_range(self, ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(r, c)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.border = copy(NO_BORDER)
                cell.fill = copy(WHITE_FILL)

    def _step_7_clear_n1_u12_preserve_n_border(self) -> None:
        ws = self._sheet("Executive Summary")
        if not ws:
            return
        self._clear_range(ws, 1, 12, column_index_from_string("N"), column_index_from_string("U"))
        for r in range(1, 13):
            c = ws[f"N{r}"]
            b = copy(c.border)
            b.left = Side(style="thin", color="FF000000")
            c.border = b
        self.log.add("Step 7: cleared N1:U12 and restored left border on N1:N12")

    def _step_8_conditional_row_collapse(self) -> None:
        ws = self._sheet("Executive Summary")
        if not ws:
            return
        l6 = self._value("Executive Summary", "L6")
        l7 = self._value("Executive Summary", "L7")

        condition_met = False
        if isinstance(l6, (int, float)) and isinstance(l7, (int, float)):
            condition_met = abs(float(l6) - float(l7)) < 1e-10
        else:
            condition_met = str(l6).strip() == str(l7).strip()

        if not condition_met:
            self.log.add(f"Step 8: condition not met (L6={l6}, L7={l7})")
            return
        ws["G7"].value = None
        ws["L7"].value = None
        ws["K5"].value = ws["K6"].value
        ws["K5"]._style = copy(ws["K6"]._style)
        ws["K6"].value = ws["K7"].value
        ws["K6"]._style = copy(ws["K7"]._style)
        ws["G7"].value = ws["G8"].value
        ws["G7"]._style = copy(ws["G8"]._style)
        ws["L7"].value = ws["L8"].value
        ws["L7"]._style = copy(ws["L8"]._style)
        ws["L7"].number_format = '0 "bps"'

        # Explicit formatting lock to match reference workbook in G5:L7.
        ws["K5"].font = copy(ws["K5"].font)
        ws["K5"].font = ws["K5"].font.copy(bold=True, underline="single")
        ws["K5"].alignment = copy(ws["K5"].alignment)
        ws["K5"].alignment = ws["K5"].alignment.copy(horizontal="right")

        ws["G7"].font = copy(ws["G7"].font)
        ws["G7"].font = ws["G7"].font.copy(bold=True, name="Cambria")
        ws["G7"].alignment = copy(ws["G7"].alignment)
        ws["G7"].alignment = ws["G7"].alignment.copy(horizontal="left")

        ws["L7"].font = copy(ws["L7"].font)
        ws["L7"].font = ws["L7"].font.copy(bold=True, name="Cambria")
        ws["L7"].alignment = copy(ws["L7"].alignment)
        ws["L7"].alignment = ws["L7"].alignment.copy(horizontal="center")
        for c in ["G8", "L8", "K7"]:
            ws[c].value = None
            ws[c].border = copy(NO_BORDER)
            ws[c].fill = copy(WHITE_FILL)
        if isinstance(ws["G6"].value, str):
            ws["G6"].value = ws["G6"].value.replace("K7", "K6")
        if isinstance(ws["L6"].value, str):
            ws["L6"].value = ws["L6"].value.replace("K7", "K6")
        if isinstance(ws["L7"].value, str):
            ws["L7"].value = ws["L7"].value.replace("L7", "L6")
        self.log.add("Step 8: row collapse executed")
    def _set_row_group_hidden(self, ws, row: int, outline: int = 1) -> None:
        rd = ws.row_dimensions[row]
        rd.hidden = True
        rd.outlineLevel = max(rd.outlineLevel or 0, outline)

    def _value_zeroish(self, v: Any) -> bool:
        if v is None:
            return True
        if isinstance(v, (int, float)):
            return v == 0
        if isinstance(v, str):
            t = v.strip()
            return t in {"", "-", " - ", "$-", "$ -"}
        return False

    def _step_9_group_hide_cash_flow_rows(self) -> None:
        ws = self._sheet("Cash Flow")
        vws = self._values_sheet("Cash Flow")
        if not ws or not vws:
            return
        grouped = []
        for row in [14, 30, 41]:
            values = [vws.cell(row, c).value for c in range(column_index_from_string("F"), column_index_from_string("Q") + 1)]
            nums = [x for x in values if isinstance(x, (int, float))]
            if nums and all(x == 0 for x in nums):
                self._set_row_group_hidden(ws, row)
                grouped.append(row)
        self.log.add(f"Step 9: grouped rows {grouped}")

    def _step_10_conditional_delete_cashflow_row(self) -> None:
        ws = self._sheet("Cash Flow")
        vws = self._values_sheet("Cash Flow")
        if not ws or not vws:
            return
        q46 = vws["Q46"].value
        q47 = vws["Q47"].value
        if q46 == q47:
            ws.delete_rows(47, 1)
            self.log.add("Step 10: deleted Cash Flow row 47")
        else:
            self.log.add("Step 10: condition not met")

    def _step_11_blue_to_black_assumptions(self) -> None:
        self._step_2_blue_to_black("Assumptions", {"0066FF", "0070C0", "00B0F0", "1F497D"})

    def _step_12_remove_comments(self, sheet: str) -> None:
        ws = self._sheet(sheet)
        if not ws:
            return
        deleted = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                if c.comment is not None:
                    c.comment = None
                    deleted += 1
        self.log.add(f"Step comments cleanup on {sheet}: removed {deleted} comments")

    def _step_42b_remove_waterfall_comments_notes(self) -> None:
        sheets = ["1-Yr Waterfall", "3-Yr Waterfall", "4-Yr Waterfall"]
        for s in sheets:
            self._step_12_remove_comments(s)
        self.log.add("Step 42b: waterfall comments/notes cleanup complete")

    def _step_13_group_hide_assumptions_dash_rows(self) -> None:
        ws = self._sheet("Assumptions")
        vws = self._values_sheet("Assumptions")
        if not ws or not vws:
            return
        grouped = []
        for row in range(49, 286):
            label = ws[f"C{row}"].value
            if label in (None, ""):
                continue
            vals = [vws.cell(row, c).value for c in range(column_index_from_string("J"), column_index_from_string("V") + 1)]
            if vals and all(self._value_zeroish(v) for v in vals):
                self._set_row_group_hidden(ws, row)
                grouped.append(row)
        self.log.add(f"Step 13: grouped {len(grouped)} Assumptions rows")

    def _step_14_hardcode_retail_tbd_sf(self) -> None:
        ws = self._sheet("Assumptions")
        vws = self._values_sheet("Assumptions")
        if not ws or not vws:
            return
        for row in range(40, 51):
            g = ws[f"G{row}"].value
            h = ws[f"H{row}"].value
            i = ws[f"I{row}"].value
            if g == "Retail" and h == "TBD" and isinstance(i, str) and i.startswith("="):
                ws[f"I{row}"].value = vws[f"I{row}"].value
                self.log.add(f"Step 14: hardcoded Assumptions!I{row}")
                return
        self.log.add("Step 14: target row not found")

    def _hardcode_formula_refs_in_sheet(self, ws_name: str, targets: tuple[str, ...]) -> int:
        ws = self._sheet(ws_name)
        vws = self._values_sheet(ws_name)
        if not ws or not vws:
            return 0
        changed = 0
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                formula = cell.value
                if isinstance(formula, str) and formula.startswith("="):
                    low = formula.lower()
                    if any(t in low for t in targets):
                        cell.value = vws.cell(r, c).value
                        changed += 1
        return changed

    def _verify_no_refs_in_sheet(self, sheet_name: str, target_sheet_name: str) -> int:
        ws = self._sheet(sheet_name)
        if not ws:
            return 0
        needle = target_sheet_name.lower()
        remaining = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and needle in cell.value.lower():
                    remaining += 1
        return remaining
    def _verify_no_refs(self, target_sheet_name: str) -> int:
        needle = target_sheet_name.lower()
        remaining = 0
        for ws in self.wb.worksheets:
            if ws.title.lower() == needle:
                continue
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("=") and needle in cell.value.lower():
                        remaining += 1
        return remaining

    def _step_15_hardcode_assumptions_external_refs(self) -> None:
        changed = self._hardcode_formula_refs_in_sheet("Assumptions", ("building program", "construction pricing"))
        remaining = self._verify_no_refs_in_sheet("Assumptions", "Building Program") + self._verify_no_refs_in_sheet("Assumptions", "Construction Pricing")
        if remaining > 0:
            raise WorkflowError(f"Step 15 failed: remaining refs count {remaining}")
        self.log.add(f"Step 15: hardcoded {changed} Assumptions external refs")

    def _step_15b_clear_total_interim_income_adjacent(self) -> None:
        ws = self._sheet("Assumptions")
        if not ws:
            self.log.add("Step 15b skipped: missing Assumptions")
            return

        hits: list[tuple[int, int]] = []
        max_row = ws.max_row
        max_col = ws.max_column
        for r in range(1, max_row):
            for c in range(1, max_col + 1):
                top = ws.cell(r, c).value
                bottom = ws.cell(r + 1, c).value
                if isinstance(top, str) and isinstance(bottom, str):
                    if top.strip().upper() == "TOTAL" and bottom.strip().upper() == "INTERIM INCOME":
                        hits.append((r, c))

        if not hits:
            self.log.add("Step 15b: condition not met")
            return

        start_cols = sorted({c + 1 for _, c in hits if c + 1 <= max_col})
        if not start_cols:
            self.log.add("Step 15b: condition found but no right-side columns to clear")
            return

        start_col = min(start_cols)
        cleared = 0
        for rr in range(1, max_row + 1):
            for cc in range(start_col, max_col + 1):
                cell = ws.cell(rr, cc)
                cell.value = None
                cell.border = copy(NO_BORDER)
                cleared += 1

        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(max_col)
        self.log.add(f"Step 15b: cleared {cleared} cells in right-side area -> {start_letter}1:{end_letter}{max_row}")

    def _step_17_non_black_non_white_to_black_development(self) -> None:
        ws = self._sheet("Development")
        if not ws:
            return
        changed = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                color = self._normalize_rgb(getattr(c.font.color, "rgb", None) if c.font and c.font.color else None)
                if color and color not in {"000000", "FFFFFF", "FFFFFE"}:
                    c.font = copy(c.font)
                    c.font = c.font.copy(color="000000")
                    changed += 1
        self.log.add(f"Step 17: changed {changed} Development text colors")

    def _step_18_hardcode_development_external_refs(self) -> None:
        changed = self._hardcode_formula_refs_in_sheet("Development", ("building program", "construction pricing"))
        remaining = self._verify_no_refs_in_sheet("Development", "Building Program") + self._verify_no_refs_in_sheet("Development", "Construction Pricing")
        if remaining > 0:
            raise WorkflowError(f"Step 18 failed: remaining refs count {remaining}")
        self.log.add(f"Step 18: hardcoded {changed} Development external refs")

    def _step_19_group_hide_development_unused_rows(self) -> None:
        ws = self._sheet("Development")
        vws = self._values_sheet("Development")
        if not ws or not vws:
            return
        grouped = []
        for row in range(1, ws.max_row + 1):
            f_label = ws[f"F{row}"].value
            if f_label in (None, ""):
                continue
            h_val = vws[f"H{row}"].value
            k_val = vws[f"K{row}"].value
            k_text = "" if k_val is None else str(k_val).strip()
            if h_val == 0 and k_text in {"-", " -"}:
                self._set_row_group_hidden(ws, row)
                grouped.append(row)
        self.log.add(f"Step 19: grouped {len(grouped)} Development rows")

    def _step_19b_clear_development_comments_description_values(self) -> None:
        ws = self._sheet("Development")
        if not ws:
            self.log.add("Step 19b skipped: missing Development")
            return

        max_row = ws.max_row
        max_col = ws.max_column
        hits: list[tuple[int, int]] = []
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip().upper() == "COMMENTS / DESCRIPTION":
                    hits.append((r, c))

        if not hits:
            self.log.add("Step 19b: condition not met")
            return

        headers_by_col: dict[int, list[int]] = {}
        for r, c in hits:
            headers_by_col.setdefault(c, []).append(r)

        cleared = 0
        for c, rows in headers_by_col.items():
            rows_sorted = sorted(rows)
            for i, hr in enumerate(rows_sorted):
                next_header = rows_sorted[i + 1] if i + 1 < len(rows_sorted) else max_row + 1
                for rr in range(hr + 1, next_header):
                    cell = ws.cell(rr, c)
                    v = cell.value
                    if v is None:
                        continue
                    if isinstance(v, str) and v.strip().upper() == "COMMENTS / DESCRIPTION":
                        continue
                    cell.value = None
                    cleared += 1

        cols = ", ".join(sorted({get_column_letter(c) for _, c in hits}))
        self.log.add(f"Step 19b: cleared {cleared} Development comment values in column(s) {cols}")

    def _step_20_group_hide_sale_proceeds_commercial(self) -> None:
        ws = self._sheet("Sale Proceeds")
        vws = self._values_sheet("Sale Proceeds")
        if not ws or not vws:
            return
        all_zero = True
        for row in range(24, 31):
            vals = [vws.cell(row, c).value for c in range(column_index_from_string("D"), column_index_from_string("M") + 1)]
            if not all(self._value_zeroish(v) for v in vals):
                all_zero = False
                break
        if all_zero:
            for row in range(24, 31):
                self._set_row_group_hidden(ws, row)
            self.log.add("Step 20: grouped Sale Proceeds rows 24:30")
        else:
            self.log.add("Step 20: condition not met")

    def _clear_to_white(self, sheet_name: str, range_address: str) -> None:
        ws = self._sheet(sheet_name)
        if not ws:
            self.log.add(f"clearToWhite skipped, missing sheet {sheet_name}")
            return
        start, end = range_address.split(":")
        start_col = column_index_from_string("".join([x for x in start if x.isalpha()]))
        start_row = int("".join([x for x in start if x.isdigit()]))
        end_col = column_index_from_string("".join([x for x in end if x.isalpha()]))
        end_row = int("".join([x for x in end if x.isdigit()]))
        self._clear_range(ws, start_row, end_row, start_col, end_col)

    def _step_28_clear_returns_exhibit(self) -> None:
        self._clear_to_white("Returns Exhibit", "J3:O37")
        self.log.add("Step 28: cleared Returns Exhibit J3:O37")

    def _step_37_clear_development_range(self) -> None:
        self._clear_to_white("Development", "L21:O31")
        self.log.add("Step 37: cleared Development L21:O31")

    def _step_38_clear_yellow_reference_cells(self) -> None:
        refs = ["'cash flow'!q46", "'executive summary'!l6", "'executive summary'!r10"]
        cleared = 0
        for sname in ["Development", "Assumptions"]:
            ws = self._sheet(sname)
            if not ws:
                continue
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for c in row:
                    formula = c.value if isinstance(c.value, str) and c.value.startswith("=") else ""
                    low = formula.lower()
                    fill = self._normalize_rgb(getattr(c.fill.fgColor, "rgb", None) if c.fill else None)
                    if fill == "FFFF00" and any(r in low for r in refs):
                        c.value = ""
                        c.fill = copy(WHITE_FILL)
                        c.border = copy(NO_BORDER)
                        cleared += 1
        self.log.add(f"Step 38: cleared {cleared} yellow reference cells")

    def _step_39_remove_non_approved_fill_colors_assumptions(self) -> None:
        ws = self._sheet("Assumptions")
        if not ws:
            return
        approved = {"002060", "DCE6F1", "FFFFFF"}
        changed = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                fill = self._normalize_rgb(getattr(c.fill.fgColor, "rgb", None) if c.fill else None)
                if fill and fill not in approved:
                    c.fill = copy(WHITE_FILL)
                    changed += 1
        self.log.add(f"Step 39: reset {changed} non-approved fills")

    def _step_40_hardcode_residential_parking_spaces(self) -> None:
        ws = self._sheet("Assumptions")
        vws = self._values_sheet("Assumptions")
        if not ws or not vws:
            raise WorkflowError("Step 40 failed: Assumptions sheet missing")
        found = None
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "residential parking spaces" in v.lower():
                    found = (r, c)
                    break
            if found:
                break
        if not found:
            raise WorkflowError("Step 40 failed: label 'Residential Parking Spaces' not found")
        rr, cc = found
        target = ws.cell(rr, cc + 1)
        if isinstance(target.value, str) and target.value.startswith("="):
            target.value = vws.cell(rr, cc + 1).value
        else:
            target.value = vws.cell(rr, cc + 1).value
        if isinstance(target.value, str) and target.value.startswith("="):
            raise WorkflowError("Step 40 failed: target cell still contains formula after hardcode")
        self.log.add(f"Step 40: hardcoded Assumptions!{get_column_letter(cc + 1)}{rr}")

    def _step_40b_hardcode_residential_parking_rent_stall(self) -> None:
        ws = self._sheet("Assumptions")
        vws = self._values_sheet("Assumptions")
        if not ws or not vws:
            raise WorkflowError("Step 40b failed: Assumptions sheet missing")

        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column):
                val = ws.cell(r, c).value
                label = str(val).strip().lower() if val is not None else ""
                if ("residential parking rent" in label and ("stall" in label or "space" in label)):
                    target = ws.cell(r, c + 1)
                    if isinstance(target.value, str) and target.value.lstrip().startswith("="):
                        target.value = vws.cell(r, c + 1).value
                        self.log.add(f"Step 40b: hardcoded Assumptions!{get_column_letter(c + 1)}{r}")
                    else:
                        self.log.add("Step 40b: target found but right cell was not a formula")
                    return

        self.log.add("Step 40b: label containing 'Residential Parking Rent' + ('Stall'/'Space') not found")
    def _step_41_copy_assumptions_range(self) -> None:
        ws = self._sheet("Assumptions")
        if not ws:
            return
        for r_off in range(0, 5):
            for c_off in range(0, 2):
                src = ws.cell(22 + r_off, column_index_from_string("W") + c_off)
                dst = ws.cell(22 + r_off, column_index_from_string("AC") + c_off)
                if isinstance(src.value, str) and src.value.startswith("="):
                    dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
                else:
                    dst.value = src.value
                dst._style = copy(src._style)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)
                dst.alignment = copy(src.alignment)
        self.log.add("Step 41: copied W22:X26 to AC22:AD26")
    def _step_42_clear_assumptions_range(self) -> None:
        self._clear_to_white("Assumptions", "W4:AA44")
        self.log.add("Step 42: cleared Assumptions W4:AA44")

    def _formula_literal(self, v: Any) -> str:
        if v is None:
            return "0"
        if isinstance(v, bool):
            return "1" if v else "0"
        if isinstance(v, (int, float)):
            return str(v)

        txt = str(v).strip()
        if txt == "":
            return "0"

        low = txt.lower()
        if low in {"-", "- ", " -", " - ", "$-", "$ -", "#value!", "#ref!", "#n/a", "#div/0!", "#num!", "#name?", "#null!"}:
            return "0"

        num_txt = txt.replace(",", "").replace("$", "").replace(" ", "")
        if num_txt in {"", "-"}:
            return "0"
        if num_txt.startswith("(") and num_txt.endswith(")"):
            num_txt = "-" + num_txt[1:-1]
        try:
            float(num_txt)
            return num_txt
        except Exception:
            pass

        esc = txt.replace('"', '""')
        return f'"{esc}"'
    def _replace_construction_pricing_refs_in_formula(self, formula: str) -> str:
        cp_vws = self._values_sheet("Construction Pricing")
        if not cp_vws:
            return formula

        pattern = re.compile(r"(?:'(?:\[[^\]]+\])?construction pricing'|(?:\[[^\]]+\])?construction pricing)!\$?([A-Z]{1,3})\$?(\d+)", re.IGNORECASE)

        def repl(match: re.Match[str]) -> str:
            col = match.group(1).upper()
            row = int(match.group(2))
            val = cp_vws[f"{col}{row}"].value
            return self._formula_literal(val)

        return pattern.sub(repl, formula)

    def _hardcode_refs_to_sheet(self, target_sheet_name: str) -> None:
        changed = 0
        needle = target_sheet_name.lower()
        for ws in self.wb.worksheets:
            if ws.title.lower() == needle:
                continue
            vws = self._values_sheet(ws.title)
            if not vws:
                continue
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(r, c)
                    if isinstance(cell.value, str) and cell.value.startswith("=") and needle in cell.value.lower():
                        # Step 24 exception: preserve Cash Flow formulas while
                        # replacing only Construction Pricing references with constants.
                        if needle == "construction pricing" and ws.title.lower() == "cash flow":
                            new_formula = self._replace_construction_pricing_refs_in_formula(cell.value)
                            cell.value = new_formula
                            changed += 1
                            continue

                        cell.value = vws.cell(r, c).value
                        changed += 1
        remaining = self._verify_no_refs(target_sheet_name)
        if remaining > 0:
            raise WorkflowError(f"Hardcode incomplete for {target_sheet_name}: {remaining} refs remain")
        self.log.add(f"Hardcoded refs to {target_sheet_name}: {changed}")
    def _safe_delete_sheet(self, sheet_name: str) -> None:
        ws = self._sheet(sheet_name)
        if not ws:
            self.log.add(f"Delete skipped: sheet '{sheet_name}' not present")
            return
        remaining = self._verify_no_refs(sheet_name)
        if remaining > 0:
            raise WorkflowError(f"Cannot delete '{sheet_name}', {remaining} references remain")
        self.wb.remove(ws)
        if self._sheet(sheet_name):
            raise WorkflowError(f"Failed to delete sheet '{sheet_name}'")
        self.log.add(f"Deleted sheet {sheet_name}")





























