from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ValidationError(RuntimeError):
    pass


class WorkflowValidator:
    def __init__(self, assertions_file: Path):
        self.assertions_file = Path(assertions_file)
        self.config: dict[str, Any] = json.loads(self.assertions_file.read_text(encoding="utf-8-sig"))

    def validate(self, workbook_file: Path) -> dict[str, Any]:
        wb = load_workbook(workbook_file, data_only=False, keep_vba=True)
        sheets = set(wb.sheetnames)
        errors: list[str] = []
        warnings: list[str] = []

        expected_removed = set(self.config.get("expected_removed_sheets", []))
        still_present = sorted(expected_removed & sheets)
        if still_present:
            errors.append(f"Removed sheets still present: {still_present}")

        if self.config.get("must_have_zero_formula_refs_to_deleted_tabs", True):
            for target in expected_removed:
                needle = target.lower()
                for ws in wb.worksheets:
                    if ws.title.lower() == needle:
                        continue
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for c in row:
                            if isinstance(c.value, str) and c.value.startswith("=") and needle in c.value.lower():
                                errors.append(f"Reference to deleted tab '{target}' at {ws.title}!{c.coordinate}")

        for check in self.config.get("spot_assertions", []):
            raw = check["cell"]
            sheet, cell = raw.split("!")
            if sheet not in wb.sheetnames:
                errors.append(f"Spot assertion sheet missing: {sheet}")
                continue
            v = wb[sheet][cell].value
            if "expected_formula" in check:
                is_formula = isinstance(v, str) and v.startswith("=")
                if is_formula != check["expected_formula"]:
                    errors.append(f"Spot assertion failed at {raw}: expected_formula={check['expected_formula']} actual={is_formula}")
            if check.get("expected_blank"):
                if v not in (None, ""):
                    errors.append(f"Spot assertion failed at {raw}: expected blank, got {v!r}")

        strict_step40 = self.config.get("validation_mode", {}).get("strict_on_step40_hardcode", True)
        if strict_step40:
            # Explicit strict rule from user: residential parking target must be hardcoded.
            # We detect by locating label text and requiring adjacent right cell non-formula.
            if "Assumptions" in wb.sheetnames:
                ws = wb["Assumptions"]
                found = None
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(r, c).value
                        if isinstance(v, str) and "residential parking spaces" in v.lower():
                            found = (r, c)
                            break
                    if found:
                        break
                if found:
                    rr, cc = found
                    right = ws.cell(rr, cc + 1).value
                    if isinstance(right, str) and right.startswith("="):
                        errors.append(
                            f"Step 40 rule failed: Assumptions!{ws.cell(rr, cc + 1).coordinate} is still formula '{right}'"
                        )
                else:
                    errors.append("Step 40 rule failed: 'Residential Parking Spaces' label not found")

        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
        }


def validate_workbook(assertions_file: Path, workbook_file: Path) -> dict[str, Any]:
    return WorkflowValidator(assertions_file).validate(workbook_file)

